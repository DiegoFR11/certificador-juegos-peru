from pathlib import Path
import re
import csv
import argparse
import logging
from copy import copy
from datetime import datetime

import fitz
from openpyxl import load_workbook


CLIENT_MARKER = "to be completed by the client"

# =============================================================================
# CERTIFICADORAS CONOCIDAS
# Para agregar una nueva empresa: añadir una entrada con su nombre interno y
# las palabras clave que aparecen en sus PDFs (en minúsculas).
#
# Ejemplo:
#   "BMM": ["bmm testlabs", "bmm.com"],
#   "ITECH": ["itech labs", "itechlabs.com"],
# =============================================================================
KNOWN_CERTIFIERS = {
    "GLI":    ["gaming laboratories international", "gaminglabs.com", "gli®"],
    "QUINEL": ["quinel"],
    # — Agregar nuevas certificadoras aquí —
}

FIELD_MAP = {
    "Game Provider": "provider",
    "Game Manufacturer": "manufacturer",
    "Game Name": "game_name",
    "Game Type": "game_type",
    "Report Reference": "report_reference",
    "Report Date": "report_date",
    "Issued by": "issued_by",
    "RNG report reference": "rng_report_reference",
    "RNG report date": "rng_report_date",
    "RNG Issued by:": "rng_issued_by",
    "RNG Issued by": "rng_issued_by",
    "RNG issued by:": "rng_issued_by",
    "RNG issued by": "rng_issued_by",
    "RNG Report Reference": "rng_report_reference",
    "RNG Report Date": "rng_report_date",
    "Sample": "sample",
    "BMM revision status": "bmm_revision_status",
    "Match with Jurisdiction in scope": "match_with_jurisdiction_in_scope",
    "General Result is PASS ": "general_result_is_pass",
    "General Result is PASS": "general_result_is_pass",
    "Unique Code": "unique_code",
    "Unique code": "unique_code",
    "Código único": "unique_code",
    "Código único de identificación": "unique_code",
}

ALWAYS_BLANK = {
    "Report date is valid?",
    "Accreditation Mark and Number",
    "Sample",
}

RE_REPORT_GAM = re.compile(r"PE_[A-Z]{3}\d+GAM\.\d+_REV\.\d+", re.I)
RE_REPORT_RNG = re.compile(r"PE_[A-Z]{3}\d+RNG\.\d+_REV\.\d+", re.I)
RE_VERSION = re.compile(r"(?:cv|v)?\d+(?:\.\d+){1,3}(?:\.?r)?|N/A", re.I)
RE_ITEM = re.compile(r"G\d{3}", re.I)

RE_GLI_REPORT_FULL = re.compile(
    r"\b[A-Z]{2}-\d{3}-[A-Z]{3}-\d{2}-\d{2,3}-\d{3}(?:\(\d+\))?\b",
    re.I,
)
RE_GLI_REPORT_SHORT = re.compile(
    r"\b[A-Z]{2}-\d{3}-[A-Z]{3}-\d{2}-\d{2,3}\b",
    re.I,
)

# Patrón genérico para referencias de informe de cualquier certificadora.
# Captura combinaciones de letras+números separadas por guiones o barras.
RE_GENERIC_REPORT = re.compile(
    r"\b[A-Z]{2,6}[-/][A-Z0-9]{2,8}[-/][A-Z0-9]{2,8}(?:[-/][A-Z0-9]{1,8})*\b",
    re.I,
)

SPANISH_MONTHS = {
    "enero": "01", "febrero": "02", "marzo": "03", "abril": "04",
    "mayo": "05", "junio": "06", "julio": "07", "agosto": "08",
    "septiembre": "09", "setiembre": "09", "octubre": "10",
    "noviembre": "11", "diciembre": "12",
}


# =============================================================================
# UTILIDADES
# =============================================================================

def clean(value):
    """Normaliza espacios y convierte None a texto vacío."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def date_to_excel(value):
    """Devuelve dd-mm-yyyy para fechas dd/mm/yyyy, dd-MMM-yy o '08 de abril de 2026'."""
    value = clean(value)

    m = re.search(r"(\d{2})/(\d{2})/(\d{4})", value)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"

    m = re.search(r"(\d{1,2})\s+de\s+([a-záéíóúñ]+)\s+de\s+(\d{4})", value, re.I)
    if m:
        day = m.group(1).zfill(2)
        month = SPANISH_MONTHS.get(m.group(2).lower())
        year = m.group(3)
        if month:
            return f"{day}-{month}-{year}"

    m = re.search(r"(\d{1,2})\s+([a-záéíóúñ]+)\s+(\d{4})", value, re.I)
    if m:
        day = m.group(1).zfill(2)
        month = SPANISH_MONTHS.get(m.group(2).lower())
        year = m.group(3)
        if month:
            return f"{day}-{month}-{year}"

    m = re.search(r"(\d{1,2})-([A-Z]{3})-(\d{2,4})", value, re.I)
    if m:
        month_map = {
            "JAN": "01", "ENE": "01", "FEB": "02", "MAR": "03", "APR": "04", "ABR": "04",
            "MAY": "05", "JUN": "06", "JUL": "07", "AUG": "08", "AGO": "08", "SEP": "09",
            "OCT": "10", "NOV": "11", "DEC": "12", "DIC": "12",
        }
        day = m.group(1).zfill(2)
        month = month_map.get(m.group(2).upper())
        year = m.group(3)
        if len(year) == 2:
            year = "20" + year
        if month:
            return f"{day}-{month}-{year}"

    return value


# =============================================================================
# DETECCIÓN DE TIPO Y CERTIFICADORA
# =============================================================================

def detect_document_type(full_text):
    """
    Identifica el tipo de documento.

    Retorna uno de:
      GLI_GAME_CERTIFICATE, QUINEL_GAME_CERTIFICATE,
      <NOMBRE>_GAME_CERTIFICATE (para certificadoras en KNOWN_CERTIFIERS),
      GENERIC_CERTIFICATE (certificado de alguna empresa no configurada),
      MINCETUR_RESOLUTION, RNG_GNA, UNKNOWN.
    """
    compact = re.sub(r"\s+", " ", full_text).lower()

    if "resolución directoral" in compact or "resolucion directoral" in compact:
        return "MINCETUR_RESOLUTION"

    for certifier, keywords in KNOWN_CERTIFIERS.items():
        if any(kw in compact for kw in keywords):
            return f"{certifier}_GAME_CERTIFICATE"

    if re.search(r"tipo de certificaci[oó]n:\s*generador de n[uú]meros aleatorios|\bGNA\b|\bRNG\b", compact, re.I):
        return "RNG_GNA"

    # Detección genérica: cualquier documento con lenguaje de certificación.
    generic_keywords = [
        "certificado de cumplimiento", "certificate of compliance",
        "compliance certificate", "test report", "certification report",
        "informe de prueba", "informe de certificación", "conformity certificate",
    ]
    if any(kw in compact for kw in generic_keywords):
        return "GENERIC_CERTIFICATE"

    return "UNKNOWN"


def try_detect_certifier(full_text):
    """
    Intenta identificar el nombre de la certificadora en documentos no configurados.
    Útil para dar contexto en el mensaje de auditoría.
    """
    compact = re.sub(r"\s+", " ", full_text)

    patterns = [
        r"emitido\s+por\s+([A-Za-zÁÉÍÓÚÑáéíóúñ0-9\s&.,\-]{4,60}?)(?:\s+el\s+\d|\s+en\s+|\.|,|\n)",
        r"issued\s+by\s+([A-Za-z0-9\s&.,\-]{4,60}?)(?:\s+on\s+\d|\.|,|\n)",
        r"prepared\s+by\s+([A-Za-z0-9\s&.,\-]{4,60}?)(?:\s+on\s+\d|\.|,|\n)",
        r"certificado\s+(?:emitido\s+)?por\s+([A-Za-zÁÉÍÓÚÑáéíóúñ0-9\s&.,\-]{4,60}?)(?:\s+|\.|,)",
    ]

    for pattern in patterns:
        m = re.search(pattern, compact, re.I)
        if m:
            name = clean(m.group(1)).strip(" ,.")
            if name and len(name) > 3:
                return name

    return ""


# =============================================================================
# EXTRACCIÓN DE ENCABEZADO — CERTIFICADORAS CONOCIDAS
# =============================================================================

def extract_gli_report_reference(compact_text):
    """Extrae la referencia principal de certificados GLI."""
    compact_text = clean(compact_text)

    label_patterns = [
        r"C[oó]digo\s+de\s+identificaci[oó]n\s+del\s+informe\s*:?\s*",
        r"N[uú]mero\s+de\s+reporte\s*:?\s*",
        r"Report\s+Reference\s*:?\s*",
        r"Report\s+Number\s*:?\s*",
    ]

    for label in label_patterns:
        m = re.search(label + f"({RE_GLI_REPORT_FULL.pattern})", compact_text, re.I)
        if m:
            return m.group(1).upper()

    candidates = RE_GLI_REPORT_FULL.findall(compact_text)
    if candidates:
        return candidates[0].upper()

    m = re.search(
        r"CERTIFICADO\s+DE\s+CUMPLIMIENTO\s+No\.?\s*"
        f"({RE_GLI_REPORT_SHORT.pattern})",
        compact_text,
        re.I,
    )
    if m:
        return m.group(1).upper()

    return ""


# =============================================================================
# EXTRACCIÓN GENÉRICA — FALLBACK PARA CERTIFICADORAS DESCONOCIDAS
# =============================================================================

def extract_generic_report_reference(compact_text):
    """
    Extrae una referencia de informe usando patrones universales.
    Se activa cuando la certificadora no está en KNOWN_CERTIFIERS.
    """
    label_patterns = [
        r"(?:report\s+(?:reference|number|no\.?)|certificate\s+no\.?|"
        r"informe\s+n[°º]?|referencia\s+del\s+informe|c[oó]digo\s+del\s+informe)\s*:?\s*"
        r"([A-Z0-9][A-Z0-9\-_/]{4,40})",
    ]

    for pattern in label_patterns:
        m = re.search(pattern, compact_text, re.I)
        if m:
            return clean(m.group(1)).upper()

    # Fallback: primer código que parezca referencia (letras+guión+números)
    candidates = RE_GENERIC_REPORT.findall(compact_text)
    for candidate in candidates:
        # Excluir fechas y versiones
        if not re.match(r"\d{2}[-/]\d{2}[-/]\d{2,4}", candidate):
            return candidate.upper()

    return ""


def extract_generic_date(compact_text):
    """Extrae la primera fecha válida que encuentre en cualquier formato."""
    date_patterns = [
        r"(?:fecha|date|issued?|emitido?)\s*:?\s*(\d{1,2}/\d{1,2}/\d{4})",
        r"(?:fecha|date|issued?|emitido?)\s*:?\s*(\d{1,2}\s+de\s+[a-záéíóúñ]+\s+de\s+\d{4})",
        r"(?:fecha|date|issued?|emitido?)\s*:?\s*(\d{1,2}\s+[a-záéíóúñ]+\s+\d{4})",
        r"(?:fecha|date|issued?|emitido?)\s*:?\s*(\d{1,2}-[A-Z]{3}-\d{2,4})",
    ]

    for pattern in date_patterns:
        m = re.search(pattern, compact_text, re.I)
        if m:
            result = date_to_excel(m.group(1))
            if result:
                return result

    # Fallback: primera fecha dd/mm/yyyy que aparezca
    m = re.search(r"\b(\d{2}/\d{2}/\d{4})\b", compact_text)
    if m:
        return date_to_excel(m.group(1))

    return ""


def extract_generic_provider(compact_text):
    """Extrae el nombre del proveedor/solicitante con patrones genéricos."""
    patterns = [
        r"(?:solicitante|applicant|cliente|client|game\s+provider)\s*:?\s*([A-Za-zÁÉÍÓÚÑáéíóúñ0-9\s&.,\-]{4,80}?)(?:\.|,|\n|$)",
        r"(?:nombre\s+y\s+datos\s+del\s+solicitante)\s*:?\s*([A-Za-zÁÉÍÓÚÑáéíóúñ0-9\s&.,\-]{4,80}?)(?:\.|,|\n|$)",
        r"(?:prepared\s+for|submitted\s+by)\s*:?\s*([A-Za-z0-9\s&.,\-]{4,80}?)(?:\.|,|\n|$)",
    ]

    for pattern in patterns:
        m = re.search(pattern, compact_text, re.I)
        if m:
            name = clean(m.group(1)).strip(" ,.")
            if name and len(name) > 3:
                return name

    return ""


def extract_generic_games(full_text):
    """
    Extracción genérica de juegos para certificadoras no configuradas.

    Busca cualquier par (nombre, versión) en el texto donde la versión
    siga el patrón estándar. No es perfecta pero es mejor que no extraer nada.
    """
    games = []
    compact = re.sub(r"\s+", " ", full_text)

    version_pat = r"(?:cv|v)?\d+(?:\.\d+){1,3}(?:\.?r)?|N/A"

    # Patrón: nombre de juego (capitalizado, 3-80 chars) seguido de versión
    row_pattern = re.compile(
        r"(?P<name>[A-ZÁÉÍÓÚÑ][A-Za-zÁÉÍÓÚÑáéíóúñ0-9+&''\-\. ]{2,80}?)\s+"
        rf"(?P<version>{version_pat})\s+"
        r"(?P<type>[A-Za-zÁÉÍÓÚÑáéíóúñ\s]{3,40}?)\s+HTML5\b",
        re.I,
    )

    seen = set()
    for idx, m in enumerate(row_pattern.finditer(compact), start=1):
        name = clean(m.group("name"))
        version = clean(m.group("version"))
        game_type = clean(m.group("type"))

        if not name or len(name) > 80:
            continue
        if re.search(r"NOMBRE|VERSI[OÓ]N|TIPO|PRODUCTO|INFORME|FECHA", name, re.I):
            continue

        key = (name.lower(), version.lower())
        if key in seen:
            continue
        seen.add(key)

        games.append({
            "item": f"GEN{idx:03d}",
            "game_name": name,
            "game_type": game_type or "Juego",
            "sample": version,
            "unique_code": "",
        })

    return games


# =============================================================================
# LECTURA DE PDF
# =============================================================================

def read_pdf_text(pdf_path):
    """Lee texto del PDF y devuelve texto completo + lista de líneas por página."""
    pdf_path = Path(pdf_path)

    if not pdf_path.exists():
        raise FileNotFoundError(f"No existe el PDF: {pdf_path}")

    pages = []
    with fitz.open(str(pdf_path)) as doc:
        if doc.is_encrypted:
            raise ValueError(f"El PDF está cifrado/protegido: {pdf_path.name}")

        for page in doc:
            lines = [clean(x) for x in page.get_text("text").splitlines() if clean(x)]
            pages.append(lines)

    full_text = "\n".join("\n".join(page_lines) for page_lines in pages)

    if not clean(full_text):
        raise ValueError(f"No se pudo extraer texto del PDF: {pdf_path.name}")

    return full_text, pages


def next_value(lines, label):
    """Busca una etiqueta y toma el valor en la misma línea o la siguiente línea no vacía."""
    pattern = re.compile(label, re.I)
    for i, line in enumerate(lines):
        if pattern.search(line):
            parts = re.split(r":", line, maxsplit=1)
            if len(parts) == 2 and clean(parts[1]):
                return clean(parts[1]).strip("•")

            for nxt in lines[i + 1:i + 8]:
                value = clean(nxt).strip("•")
                if value:
                    return value
    return ""


# =============================================================================
# EXTRACCIÓN DE ENCABEZADO
# =============================================================================

def extract_expected_count(full_text):
    """Extrae conteo esperado desde frases como 'Tipo de Producto: ... (30 juegos)'."""
    compact = re.sub(r"\s+", " ", full_text)
    m = re.search(r"Tipo de Producto:.*?\((\d+)\s*juego", compact, re.I)
    if m:
        return int(m.group(1))
    return None


def extract_header(full_text, pages):
    """
    Extrae el encabezado del certificado.

    Para certificadoras en KNOWN_CERTIFIERS usa extractores específicos.
    Para documentos GENERIC_CERTIFICATE o UNKNOWN aplica fallbacks genéricos
    e intenta extraer los campos clave con patrones universales.
    """
    first = pages[0] if pages else []
    compact = re.sub(r"\s+", " ", full_text)
    doc_type = detect_document_type(full_text)

    report_reference = ""

    # ── Extracción específica por certificadora ──────────────────────────────
    m = re.search(
        r"CERTIFICADO DE CUMPLIMIENTO\s*N[°º]?\s*(PE_[A-Z]{3}\d+GAM\.\d+_REV\.\d+)",
        compact, re.I,
    )
    if not m:
        m = re.search(r"ID del Informe:\s*(PE_[A-Z]{3}\d+GAM\.\d+_REV\.\d+)", compact, re.I)
    if not m:
        m = RE_REPORT_GAM.search(compact)
    if m:
        report_reference = m.group(1).upper()

    if not report_reference:
        report_reference = extract_gli_report_reference(compact)

    # ── Fallback genérico para certificadoras desconocidas ───────────────────
    if not report_reference and doc_type in ("GENERIC_CERTIFICATE", "UNKNOWN"):
        report_reference = extract_generic_report_reference(compact)

    # ── Fecha ────────────────────────────────────────────────────────────────
    report_date = ""

    m = re.search(r"Fecha de emisión:\s*(\d{2}/\d{2}/\d{4})", compact, re.I)
    if m:
        report_date = date_to_excel(m.group(1))

    if not report_date:
        m = re.search(r"Fecha:\s*(\d{1,2}\s+de\s+[a-záéíóúñ]+\s+de\s+\d{4})", compact, re.I)
        if m:
            report_date = date_to_excel(m.group(1))

    if not report_date:
        m = re.search(r"Fecha:\s*(\d{1,2}\s+[a-záéíóúñ]+\s+\d{4})", compact, re.I)
        if m:
            report_date = date_to_excel(m.group(1))

    if not report_date:
        for i, line in enumerate(first):
            if clean(line).lower().startswith("fecha"):
                for nxt in first[i:i + 5]:
                    if re.search(r"\d{2}/\d{2}/\d{4}", nxt):
                        report_date = date_to_excel(nxt)
                        break
            if report_date:
                break

    if not report_date and doc_type in ("GENERIC_CERTIFICATE", "UNKNOWN"):
        report_date = extract_generic_date(compact)

    # ── Proveedor / fabricante ───────────────────────────────────────────────
    provider = next_value(first, r"Nombre y datos del solicitante")
    manufacturer = next_value(first, r"Nombre y datos del Fabricante") or provider

    if not provider and doc_type in ("GENERIC_CERTIFICATE", "UNKNOWN"):
        provider = extract_generic_provider(compact)
        manufacturer = provider

    # ── Datos RNG ────────────────────────────────────────────────────────────
    rng_report_reference = ""
    rng_report_date = ""
    rng_issued_by = ""

    m = re.search(
        r"(PE_[A-Z]{3}\d+RNG\.\d+_REV\.\d+)\s+emitido\s+por\s+([^\.]+?)\s+el\s+(\d{2}/\d{2}/\d{4})",
        compact, re.I,
    )
    if m:
        rng_report_reference = m.group(1).upper()
        rng_issued_by = "QUINEL Ltd" if "quinel" in m.group(2).lower() else clean(m.group(2))
        rng_report_date = date_to_excel(m.group(3))
    else:
        m = RE_REPORT_RNG.search(compact)
        if m:
            rng_report_reference = m.group(0).upper()
            rng_issued_by = "QUINEL Ltd" if "QUINEL" in full_text else ""

    if not rng_report_reference:
        date_pattern = (
            r"\d{1,2}/\d{1,2}/\d{4}"
            r"|\d{1,2}\s+(?:de\s+)?[a-záéíóúñ]+\s+(?:de\s+)?\d{4}"
            r"|\d{1,2}-[A-Z]{3}-\d{2,4}"
        )
        m = re.search(
            r"Certificado\s+de\s+Cumplimiento\s+No\.?\s+"
            rf"({RE_GLI_REPORT_FULL.pattern})"
            r"\s+emitido\s+por\s+(.+?)\s+el\s+"
            rf"({date_pattern})",
            compact, re.I,
        )
        if m:
            rng_report_reference = m.group(1).upper()
            issuer = clean(m.group(2))
            rng_issued_by = "GLI" if "gli" in issuer.lower() else issuer
            rng_report_date = date_to_excel(m.group(3))

    # ── issued_by ────────────────────────────────────────────────────────────
    if "GLI_GAME_CERTIFICATE" in doc_type:
        issued_by = "GLI"
    elif "QUINEL_GAME_CERTIFICATE" in doc_type:
        issued_by = "QUINEL Ltd"
    elif "GAMING LABORATORIES INTERNATIONAL" in full_text.upper():
        issued_by = "GLI"
    elif doc_type in ("GENERIC_CERTIFICATE", "UNKNOWN"):
        issued_by = try_detect_certifier(compact)
    else:
        issued_by = ""

    jurisdiction = "YES" if re.search(r"Jurisdicci[oó]n:?\s*Per[uú]", compact, re.I) else ""
    result_pass = "YES" if re.search(r"Conclusi[oó]n:?\s*CUMPLE", compact, re.I) else ""

    return {
        "provider": provider,
        "manufacturer": manufacturer,
        "report_reference": report_reference,
        "report_date": report_date,
        "issued_by": issued_by,
        "rng_report_reference": rng_report_reference,
        "rng_report_date": rng_report_date,
        "rng_issued_by": rng_issued_by,
        "bmm_revision_status": "",
        "match_with_jurisdiction_in_scope": jurisdiction,
        "general_result_is_pass": result_pass,
        "expected_games": extract_expected_count(full_text),
        "document_type": doc_type,
    }


# =============================================================================
# EXTRACCIÓN DE JUEGOS
# =============================================================================

def get_summary_text(full_text):
    """Limita la extracción de juegos a A.1.1 Resumen, antes de A.1.2 Archivos Críticos."""
    compact = re.sub(r"\s+", " ", full_text)
    summary_text = compact

    start_match = re.search(r"A\.1\.1\.?\s*Resumen", compact, re.I)
    if start_match:
        summary_text = compact[start_match.start():]

    end_match = re.search(r"A\.1\.2\.?\s*Archivos Cr[ií]ticos", summary_text, re.I)
    if end_match:
        summary_text = summary_text[:end_match.start()]

    return summary_text


def normalize_game_name(name):
    name = clean(name)
    name = re.sub(r"^ITEM\s+NOMBRE DEL\s+PRODUCTO\s+", "", name, flags=re.I).strip()
    name = re.sub(r"^ITEM\s+NOMBRE DEL PRODUCTO\s+", "", name, flags=re.I).strip()
    name = re.sub(r"\s+Código único.*$", "", name, flags=re.I).strip()
    name = re.sub(r"\s+VERSI[OÓ]N.*$", "", name, flags=re.I).strip()
    return name


def normalize_game_type(game_type):
    game_type = clean(game_type)
    game_type = re.sub(r"\s+MEDIOS SOPORTADOS.*$", "", game_type, flags=re.I).strip()
    return game_type


def split_joined_unique_code(text):
    """Detecta códigos únicos incluso si el PDF los parte en dos líneas/tokens."""
    value = clean(text)

    m = re.search(r"([A-Za-z0-9]{12,}_[0-9.]+)$", value)
    if m:
        return value[:m.start()].strip(), m.group(1)

    m = re.search(r"([A-Za-z0-9]{8,})\s+([A-Za-z0-9]{6,}_[0-9.]+)$", value)
    if m:
        unique_code = f"{m.group(1)}{m.group(2)}"
        return value[:m.start()].strip(), unique_code

    return value, ""


def is_version(value):
    return bool(RE_VERSION.fullmatch(clean(value)))


def is_unique_code(value):
    value = clean(value)
    return bool(
        re.fullmatch(r"[A-Za-z0-9]{12,}_[0-9.]+", value)
        or re.fullmatch(r"[A-Za-z0-9]{8,}\s+[A-Za-z0-9]{6,}_[0-9.]+", value)
    )


def extract_games_from_compact_summary(summary_text):
    """Extrae juegos desde tablas convertidas a texto continuo (QUINEL)."""
    games = []

    pattern = re.compile(
        r"\b(G\d{3})\s+"
        r"(.+?)\s+"
        r"(\d+(?:\.\d+){1,3})\s+"
        r"(.+?)\s+"
        r"HTML5",
        re.I,
    )

    for m in pattern.finditer(summary_text):
        item = clean(m.group(1)).upper()
        raw_name = normalize_game_name(m.group(2))
        version = clean(m.group(3))
        game_type = normalize_game_type(m.group(4))

        name, unique_code = split_joined_unique_code(raw_name)

        games.append({
            "item": item,
            "game_name": name,
            "game_type": game_type or "Juegos de Línea",
            "sample": version,
            "unique_code": unique_code,
        })

    return games


def extract_games_from_gli_summary(summary_text):
    """Extrae juegos desde certificados GLI con columna CÓDIGO ÚNICO."""
    games = []

    header = re.search(
        r"NOMBRE\s+DEL\s+PRODUCTO\s+C[OÓ]DIGO\s+[UÚ]NICO\s+VERSI[OÓ]N\s+TIPO\s+DE\s+JUEGO\s+MEDIOS\s+SOPORTADOS",
        summary_text, re.I,
    )
    if not header:
        return games

    table_text = summary_text[header.end():]
    table_text = re.split(r"\bLa plataforma tecnol[oó]gica\b|\bA\.1\.2\b", table_text, flags=re.I)[0]
    table_text = clean(table_text)

    row_pattern = re.compile(
        r"(?P<name>[A-ZÁÉÍÓÚÑ0-9][A-Za-zÁÉÍÓÚÑáéíóúñ0-9+&''\- .]{1,120}?)\s+"
        r"(?P<unique>[A-Za-z0-9]+(?:[_-][A-Za-z0-9]+)+)\s+"
        r"(?P<version>(?:cv|v)?\d+(?:\.\d+)+)\s+"
        r"(?P<type>.+?)\s+HTML5\b",
        re.I,
    )

    for idx, m in enumerate(row_pattern.finditer(table_text), start=1):
        name = normalize_game_name(m.group("name"))
        unique_code = clean(m.group("unique"))
        version = clean(m.group("version"))
        game_type = normalize_game_type(m.group("type"))

        if not name or len(name) > 120:
            continue

        games.append({
            "item": f"GLI{idx:03d}",
            "game_name": name,
            "game_type": game_type or "Juego de Tragamonedas",
            "sample": version,
            "unique_code": unique_code,
        })

    return games


def extract_games_from_gli_no_code_summary(summary_text):
    """Extrae juegos desde certificados GLI sin CÓDIGO ÚNICO (ej. Amusnet)."""
    games = []

    header = re.search(
        r"NOMBRE\s+DEL\s+PRODUCTO\s+VERSI[OÓ]N\s+TIPO\s+DE\s+JUEGO\s+MEDIOS\s+SOPORTADO\s*S",
        summary_text, re.I,
    )
    if not header:
        return games

    table_text = summary_text[header.end():]
    table_text = re.split(
        r"\bLa informaci[oó]n de la plataforma\b|\bA\.1\.2\b|\bArchivos Cr[ií]ticos\b",
        table_text, flags=re.I,
    )[0]
    table_text = clean(table_text)

    version_pattern = r"(?:cv|v)?\d+(?:\.\d+){1,3}(?:\.?r)?|N/A"

    row_pattern = re.compile(
        r"(?P<name>[A-ZÁÉÍÓÚÑ0-9][A-Za-zÁÉÍÓÚÑáéíóúñ0-9+&''\-., ]{1,120}?)\s+"
        rf"(?P<version>{version_pattern})\s+"
        r"(?P<type>Tragamonedas|Ruleta|Funci[oó]n\s+progresiva|Juego\s+de\s+[^H]+?)\s+HTML5\b",
        re.I,
    )

    for idx, m in enumerate(row_pattern.finditer(table_text), start=1):
        name = normalize_game_name(m.group("name"))
        name = re.sub(r"^.*?GAMINGLABS\.com\s+", "", name, flags=re.I).strip()
        version = clean(m.group("version"))
        game_type = normalize_game_type(m.group("type"))

        if not name or len(name) > 120:
            continue
        if re.search(r"NOMBRE\s+DEL\s+PRODUCTO|VERSI[OÓ]N|TIPO\s+DE\s+JUEGO", name, re.I):
            continue

        games.append({
            "item": f"GLI{idx:03d}",
            "game_name": name,
            "game_type": game_type or "Tragamonedas",
            "sample": version,
            "unique_code": "",
        })

    return games


def extract_games_from_lines(pages):
    """Fallback para PDFs donde la tabla queda partida en líneas."""
    games = []

    limited_pages = []
    stop = False
    for lines in pages:
        current = []
        for line in lines:
            if re.search(r"A\.1\.2|Archivos Cr[ií]ticos", line, re.I):
                stop = True
                break
            current.append(line)
        limited_pages.append(current)
        if stop:
            break

    for lines in limited_pages:
        i = 0
        while i < len(lines):
            value = clean(lines[i])

            if re.fullmatch(r"G\d{3}", value):
                item = value.upper()
                ptr = i + 1
                name_parts = []
                unique_code_parts = []

                while ptr < len(lines):
                    current = clean(lines[ptr])

                    if is_version(current):
                        break
                    if is_unique_code(current):
                        unique_code_parts.append(current)
                        ptr += 1
                        continue
                    if re.fullmatch(r"G\d{3}", current):
                        break
                    if re.search(r"ITEM|VERSI[OÓ]N|TIPO DE|MEDIOS SOPORTADOS|Código único", current, re.I):
                        ptr += 1
                        continue
                    if re.fullmatch(r"[A-Za-z0-9]{8,}", current) or re.fullmatch(r"[A-Za-z0-9]{6,}_[0-9.]+", current):
                        unique_code_parts.append(current)
                    else:
                        name_parts.append(current)

                    ptr += 1

                if ptr < len(lines) and is_version(lines[ptr]):
                    version = clean(lines[ptr])
                    ptr += 1

                    game_type_parts = []
                    while ptr < len(lines):
                        current = clean(lines[ptr])
                        if re.fullmatch(r"G\d{3}", current):
                            break
                        if re.search(r"HTML5|MEDIOS|SOPORTADOS|A\.1\.2", current, re.I):
                            break
                        game_type_parts.append(current)
                        ptr += 1

                    game_type = normalize_game_type(" ".join(game_type_parts))
                    name = normalize_game_name(" ".join(name_parts))
                    unique_code = "".join(unique_code_parts).replace(" ", "")

                    if name and version:
                        games.append({
                            "item": item,
                            "game_name": name,
                            "game_type": game_type or "Juegos de Línea",
                            "sample": version,
                            "unique_code": unique_code,
                        })

                    i = ptr
                    continue

            i += 1

    return games


def normalize_for_key(value):
    """Normaliza texto para comparar juegos entre distintos extractores."""
    value = clean(value).lower()
    return re.sub(r"\s+", " ", value).strip()


def game_score(game):
    """Puntaje para conservar el registro más completo."""
    score = 0
    if clean(game.get("game_name")):
        score += 2
    if clean(game.get("game_type")):
        score += 1
    if clean(game.get("sample")):
        score += 1
    if clean(game.get("unique_code")):
        score += 3
    if clean(game.get("item")):
        score += 1
    return score


def game_identity_keys(game):
    """Genera llaves de identidad para detectar duplicados reales."""
    item = normalize_for_key(game.get("item"))
    name = normalize_for_key(game.get("game_name"))
    sample = normalize_for_key(game.get("sample"))
    unique_code = normalize_for_key(game.get("unique_code"))

    keys = []
    if unique_code:
        keys.append(("unique_code", unique_code))
    if name and sample:
        keys.append(("name_sample", name, sample))
    if name:
        keys.append(("name", name))
    if item:
        keys.append(("item", item))
    return keys


def dedupe_games(games):
    """Deduplica juegos por identidad lógica, priorizando el registro más completo."""
    key_to_game = {}

    for game in games:
        keys = game_identity_keys(game)
        if not keys:
            continue

        existing_games = [key_to_game[key] for key in keys if key in key_to_game]

        if existing_games:
            current_best = max(existing_games, key=game_score)
            best = game if game_score(game) > game_score(current_best) else current_best
        else:
            best = game

        for key in keys:
            key_to_game[key] = best

    unique = []
    seen_ids = set()

    for game in key_to_game.values():
        obj_id = id(game)
        if obj_id not in seen_ids:
            unique.append(game)
            seen_ids.add(obj_id)

    return sorted(
        unique,
        key=lambda g: normalize_for_key(g.get("item")) or normalize_for_key(g.get("game_name")),
    )


def extract_games(full_text, pages):
    """
    Orquesta todos los extractores de juegos.

    Para certificadoras conocidas (GLI, QUINEL) usa extractores específicos.
    Para documentos GENERIC_CERTIFICATE o UNKNOWN intenta el extractor genérico
    como último recurso.
    """
    summary_text = get_summary_text(full_text)
    doc_type = detect_document_type(full_text)

    games = []
    games.extend(extract_games_from_compact_summary(summary_text))
    games.extend(extract_games_from_lines(pages))
    games.extend(extract_games_from_gli_summary(summary_text))
    games.extend(extract_games_from_gli_no_code_summary(summary_text))

    # Para documentos desconocidos, intentar extracción genérica si los
    # extractores específicos no encontraron nada.
    if not games and doc_type in ("GENERIC_CERTIFICATE", "UNKNOWN"):
        games.extend(extract_generic_games(full_text))

    cleaned_games = []
    for game in dedupe_games(games):
        name = clean(game.get("game_name"))
        sample = clean(game.get("sample"))

        if not name or not sample:
            continue
        if len(name) > 120:
            logging.warning("Juego descartado por nombre demasiado largo: %s", name)
            continue

        cleaned_games.append({
            "game_name": name,
            "game_type": clean(game.get("game_type")) or "Juegos de Línea",
            "sample": sample,
            "unique_code": clean(game.get("unique_code")),
        })

    return cleaned_games


# =============================================================================
# ESCRITURA EN EXCEL
# =============================================================================

def merged_value(ws, row, col):
    cell = ws.cell(row, col)
    if cell.value is not None:
        return cell.value

    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return ws.cell(merged.min_row, merged.min_col).value

    return None


def find_sheet(wb):
    if "B2B table" in wb.sheetnames:
        return wb["B2B table"]

    for ws in wb.worksheets:
        for r in range(1, min(ws.max_row, 30) + 1):
            values = [clean(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
            if "Game Provider" in values and "Game Name" in values:
                return ws

    raise ValueError("No encontré la hoja B2B table ni una hoja con los encabezados esperados.")


def find_header_row(ws):
    for r in range(1, min(ws.max_row, 40) + 1):
        values = [clean(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if "Game Provider" in values and "Game Name" in values and "Report Reference" in values:
            return r

    raise ValueError("No encontré la fila de encabezados.")


def copy_row_format(ws, source_row, target_row):
    for col in range(1, ws.max_column + 1):
        source_cell = ws.cell(source_row, col)
        target_cell = ws.cell(target_row, col)

        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)

        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)

    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def fill_excel(template_path, rows, output_path):
    template_path = Path(template_path)
    output_path = Path(output_path)

    if not template_path.exists():
        raise FileNotFoundError(f"No existe la plantilla: {template_path}")

    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(template_path)
    ws = find_sheet(wb)
    header_row = find_header_row(ws)
    marker_row = header_row - 1
    data_row = header_row + 1

    fillable_columns = {}

    for col in range(1, ws.max_column + 1):
        header = clean(ws.cell(header_row, col).value)
        marker = clean(merged_value(ws, marker_row, col)).lower()

        if marker == CLIENT_MARKER or header == "Sample":
            fillable_columns[col] = header

    if not fillable_columns:
        raise ValueError("No encontré columnas editables en la plantilla.")

    style_source_row = data_row

    if ws.max_row > data_row:
        ws.delete_rows(data_row + 1, ws.max_row - data_row)

    for col in range(1, ws.max_column + 1):
        ws.cell(data_row, col).value = None

    rows_to_create = max(len(rows), 1)

    for idx in range(rows_to_create):
        excel_row = data_row + idx
        copy_row_format(ws, style_source_row, excel_row)
        for col in range(1, ws.max_column + 1):
            ws.cell(excel_row, col).value = None

    for idx, row in enumerate(rows):
        excel_row = data_row + idx
        for col, header in fillable_columns.items():
            if header in ALWAYS_BLANK:
                ws.cell(excel_row, col).value = None
                continue
            field = FIELD_MAP.get(header)
            if not field:
                continue
            value = row.get(field, "")
            ws.cell(excel_row, col).value = value or None

    wb.save(output_path)


# =============================================================================
# PROCESAMIENTO PRINCIPAL
# =============================================================================

def build_rows_for_pdf(pdf):
    """
    Procesa un PDF y construye las filas para el Excel.

    El mensaje de auditoría es descriptivo: indica la certificadora detectada
    y lista exactamente qué campos no se pudieron extraer.
    """
    pdf_path = Path(pdf)
    full_text, pages = read_pdf_text(pdf_path)
    header = extract_header(full_text, pages)
    games = extract_games(full_text, pages)

    rows = []
    for game in games:
        row = {}
        row.update(header)
        row.update(game)
        row.pop("expected_games", None)
        row.pop("document_type", None)
        rows.append(row)

    expected = header.get("expected_games")
    extracted = len(games)
    doc_type = header.get("document_type", "UNKNOWN")

    # ── Diagnóstico detallado ─────────────────────────────────────────────────
    missing_fields = []
    if not header.get("report_reference"):
        missing_fields.append("Referencia del informe")
    if not header.get("report_date"):
        missing_fields.append("Fecha del informe")
    if not header.get("provider"):
        missing_fields.append("Proveedor / Game Provider")
    if not header.get("issued_by"):
        missing_fields.append("Certificadora (Issued by)")

    certifier_hint = ""
    if doc_type in ("GENERIC_CERTIFICATE", "UNKNOWN"):
        detected = try_detect_certifier(re.sub(r"\s+", " ", full_text))
        if detected:
            certifier_hint = (
                f" Empresa detectada: '{detected}'. "
                f"Agrégala en KNOWN_CERTIFIERS en generar_excel.py para extracción completa."
            )
        else:
            certifier_hint = (
                " No se identificó la certificadora. "
                "Verifica que el PDF corresponda a un certificado de cumplimiento."
            )

    if extracted == 0:
        status = "REVISAR"
        message = f"No se extrajeron juegos. Tipo detectado: {doc_type}.{certifier_hint}"
    elif expected is not None and expected != extracted:
        status = "REVISAR"
        message = f"Esperados {expected}, extraídos {extracted}."
    elif missing_fields:
        status = "REVISAR"
        message = "Campos no extraídos: " + ", ".join(missing_fields) + "." + certifier_hint
    else:
        status = "OK"
        message = ""

    return rows, {
        "pdf": pdf_path.name,
        "document_type": doc_type,
        "report_reference": header.get("report_reference", ""),
        "expected_games": expected if expected is not None else "",
        "extracted_games": extracted,
        "status": status,
        "message": message,
    }


def write_audit_csv(audit_rows, audit_path):
    audit_path = Path(audit_path)
    audit_path.parent.mkdir(parents=True, exist_ok=True)

    columns = ["pdf", "document_type", "report_reference", "expected_games", "extracted_games", "status", "message"]

    with audit_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=columns)
        writer.writeheader()
        for row in audit_rows:
            writer.writerow({col: row.get(col, "") for col in columns})


def resolve_pdfs(pdf_args, pdf_dir):
    pdfs = []

    for pdf in pdf_args or []:
        pdfs.append(Path(pdf))

    if pdf_dir:
        pdf_dir = Path(pdf_dir)
        if not pdf_dir.exists():
            raise FileNotFoundError(f"No existe la carpeta de PDFs: {pdf_dir}")
        pdfs.extend(sorted(pdf_dir.glob("*.pdf")))

    seen = set()
    unique_pdfs = []
    for pdf in pdfs:
        key = str(pdf.resolve())
        if key not in seen:
            unique_pdfs.append(pdf)
            seen.add(key)

    if not unique_pdfs:
        raise ValueError("No se recibió ningún PDF. Usa --pdf o --pdf-dir.")

    return unique_pdfs


def process(template, pdfs, output, audit=None, strict=False):
    all_rows = []
    audit_rows = []

    for pdf in pdfs:
        try:
            rows, audit_row = build_rows_for_pdf(pdf)
            all_rows.extend(rows)
            audit_rows.append(audit_row)

            logging.info(
                "%s: %s juegos extraídos. Estado: %s",
                audit_row["pdf"],
                audit_row["extracted_games"],
                audit_row["status"],
            )

            if strict and audit_row["status"] != "OK":
                raise ValueError(audit_row["message"])

        except Exception as exc:
            audit_row = {
                "pdf": Path(pdf).name,
                "document_type": "ERROR",
                "report_reference": "",
                "expected_games": "",
                "extracted_games": 0,
                "status": "ERROR",
                "message": str(exc),
            }
            audit_rows.append(audit_row)
            logging.exception("Error procesando %s", pdf)

            if strict:
                raise

    fill_excel(template, all_rows, output)

    if audit:
        write_audit_csv(audit_rows, audit)

    logging.info("Excel generado: %s", output)
    logging.info("Total de filas: %s", len(all_rows))

    revisar = [row for row in audit_rows if row["status"] != "OK"]
    if revisar:
        logging.warning("Hay %s PDF(s) con estado REVISAR/ERROR. Consulta el audit CSV.", len(revisar))

    return audit_rows


def parse_args():
    parser = argparse.ArgumentParser(
        description="Genera Excel B2B de integración de juegos a partir de certificados PDF para Perú."
    )
    parser.add_argument("--template", required=True, help="Ruta de la plantilla Excel.")
    parser.add_argument("--pdf", action="append", help="Ruta de un PDF. Se puede repetir.")
    parser.add_argument("--pdf-dir", help="Carpeta con PDFs. Procesa todos los *.pdf.")
    parser.add_argument("--output", required=True, help="Ruta del Excel generado.")
    parser.add_argument("--audit", help="Ruta opcional para CSV de auditoría.")
    parser.add_argument(
        "--strict",
        action="store_true",
        help="Detiene el proceso si un PDF falla o si el conteo esperado no coincide.",
    )
    parser.add_argument(
        "--log-level",
        default="INFO",
        choices=["DEBUG", "INFO", "WARNING", "ERROR"],
        help="Nivel de log.",
    )
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()

    logging.basicConfig(
        level=getattr(logging, args.log_level),
        format="%(levelname)s | %(message)s",
    )

    pdfs = resolve_pdfs(args.pdf, args.pdf_dir)

    process(
        template=args.template,
        pdfs=pdfs,
        output=args.output,
        audit=args.audit,
        strict=args.strict,
    )
