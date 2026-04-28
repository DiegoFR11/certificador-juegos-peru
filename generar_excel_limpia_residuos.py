from pathlib import Path
import re
import csv
import argparse
import logging
from copy import copy
from datetime import datetime

import fitz
from openpyxl import load_workbook


CLIENT_MARKER = "to be completed by the client"

FIELD_MAP = {
    "Game Provider": "provider",
    "Game Manufacturer": "manufacturer",
    "Game Name": "game_name",
    "Game Type": "game_type",
    "Report Reference": "report_reference",
    "Report Date": "report_date",
    "Issued by": "issued_by",
    "RNG report reference": "rng_report_reference",
    "RNG report date": "rng_report_date",
    "RNG Issued by:": "rng_issued_by",
    "RNG Issued by": "rng_issued_by",
    "Sample": "sample",
    "BMM revision status": "bmm_revision_status",
    "Match with Jurisdiction in scope": "match_with_jurisdiction_in_scope",
    "General Result is PASS ": "general_result_is_pass",
    "General Result is PASS": "general_result_is_pass",
    # Posibles nombres si la plantilla llega a incorporar el código único.
    "Unique Code": "unique_code",
    "Unique code": "unique_code",
    "Código único": "unique_code",
    "Código único de identificación": "unique_code",
}

ALWAYS_BLANK = {
    "Report date is valid?",
    "Accreditation Mark and Number",
    "Sample",
}

RE_REPORT_GAM = re.compile(r"PE_[A-Z]{3}\d+GAM\.\d+_REV\.\d+", re.I)
RE_REPORT_RNG = re.compile(r"PE_[A-Z]{3}\d+RNG\.\d+_REV\.\d+", re.I)
RE_VERSION = re.compile(r"\d+(?:\.\d+){1,3}")
RE_ITEM = re.compile(r"G\d{3}", re.I)


def clean(value):
    """Normaliza espacios y convierte None a texto vacío."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def date_to_excel(value):
    """Convierte fechas dd/mm/yyyy a dd-mm-yyyy. Si no encuentra fecha, devuelve texto limpio."""
    m = re.search(r"(\d{2})/(\d{2})/(\d{4})", value or "")
    if not m:
        return clean(value)
    return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"


def split_joined_unique_code(text):
    """
    Detecta códigos únicos incluso si el PDF los parte en dos líneas/tokens.

    Ejemplo:
    66436d3fc069e70 0017e663e_961 -> 66436d3fc069e700017e663e_961
    """
    value = clean(text)

    # Caso 1: código completo en un solo token.
    m = re.search(r"([A-Za-z0-9]{12,}_[0-9.]+)$", value)
    if m:
        return value[:m.start()].strip(), m.group(1)

    # Caso 2: código partido en dos tokens antes de la versión.
    # Se permite una primera parte larga y una segunda parte con _.
    m = re.search(r"([A-Za-z0-9]{8,})\s+([A-Za-z0-9]{6,}_[0-9.]+)$", value)
    if m:
        unique_code = f"{m.group(1)}{m.group(2)}"
        return value[:m.start()].strip(), unique_code

    return value, ""


def is_version(value):
    return bool(re.fullmatch(r"\d+(?:\.\d+){1,3}", clean(value)))


def is_unique_code(value):
    value = clean(value)
    return bool(
        re.fullmatch(r"[A-Za-z0-9]{12,}_[0-9.]+", value)
        or re.fullmatch(r"[A-Za-z0-9]{8,}\s+[A-Za-z0-9]{6,}_[0-9.]+", value)
    )


def read_pdf_text(pdf_path):
    """Lee texto del PDF y devuelve texto completo + lista de líneas por página."""
    pdf_path = Path(pdf_path)

    if not pdf_path.exists():
        raise FileNotFoundError(f"No existe el PDF: {pdf_path}")

    pages = []
    with fitz.open(str(pdf_path)) as doc:
        if doc.is_encrypted:
            raise ValueError(f"El PDF está cifrado/protegido: {pdf_path.name}")

        for page in doc:
            lines = [clean(x) for x in page.get_text("text").splitlines() if clean(x)]
            pages.append(lines)

    full_text = "\n".join("\n".join(page_lines) for page_lines in pages)

    if not clean(full_text):
        raise ValueError(f"No se pudo extraer texto del PDF: {pdf_path.name}")

    return full_text, pages


def next_value(lines, label):
    """Busca una etiqueta y toma la siguiente línea no vacía."""
    pattern = re.compile(label, re.I)
    for i, line in enumerate(lines):
        if pattern.search(line):
            for nxt in lines[i + 1:i + 8]:
                if clean(nxt):
                    return clean(nxt)
    return ""


def extract_expected_count(full_text):
    """
    Extrae el conteo esperado desde frases como:
    Tipo de Producto: ... (30 juegos)
    Tipo de Producto: ... (1 juego)
    """
    compact = re.sub(r"\s+", " ", full_text)
    m = re.search(r"Tipo de Producto:.*?\((\d+)\s*juego", compact, re.I)
    if m:
        return int(m.group(1))
    return None


def extract_header(full_text, pages):
    first = pages[0] if pages else []
    compact = re.sub(r"\s+", " ", full_text)

    report_reference = ""
    m = re.search(
        r"CERTIFICADO DE CUMPLIMIENTO\s*N[°º]?\s*(PE_[A-Z]{3}\d+GAM\.\d+_REV\.\d+)",
        compact,
        re.I,
    )
    if not m:
        m = re.search(r"ID del Informe:\s*(PE_[A-Z]{3}\d+GAM\.\d+_REV\.\d+)", compact, re.I)
    if not m:
        m = RE_REPORT_GAM.search(compact)
    if m:
        report_reference = m.group(1).upper()

    report_date = ""
    # Prioridad 1: Fecha de emisión.
    m = re.search(r"Fecha de emisión:\s*(\d{2}/\d{2}/\d{4})", compact, re.I)
    if m:
        report_date = date_to_excel(m.group(1))
    else:
        # Prioridad 2: etiqueta Fecha en primera página.
        for i, line in enumerate(first):
            if clean(line).lower().startswith("fecha"):
                for nxt in first[i:i + 5]:
                    if re.search(r"\d{2}/\d{2}/\d{4}", nxt):
                        report_date = date_to_excel(nxt)
                        break
            if report_date:
                break

    provider = next_value(first, r"Nombre y datos del solicitante")
    manufacturer = next_value(first, r"Nombre y datos del Fabricante") or provider

    rng_report_reference = ""
    rng_report_date = ""
    rng_issued_by = ""

    # Caso con emisor y fecha.
    m = re.search(
        r"(PE_[A-Z]{3}\d+RNG\.\d+_REV\.\d+)\s+emitido\s+por\s+([^\.]+?)\s+el\s+(\d{2}/\d{2}/\d{4})",
        compact,
        re.I,
    )
    if m:
        rng_report_reference = m.group(1).upper()
        rng_issued_by = "QUINEL Ltd" if "quinel" in m.group(2).lower() else clean(m.group(2))
        rng_report_date = date_to_excel(m.group(3))
    else:
        # Caso frecuente en los certificados: Hacer referencia informe con ID "PE_...RNG..."
        m = RE_REPORT_RNG.search(compact)
        if m:
            rng_report_reference = m.group(0).upper()
            rng_issued_by = "QUINEL Ltd" if "QUINEL" in full_text else ""

    jurisdiction = "YES" if re.search(r"Jurisdicción\s+Per[uú]", compact, re.I) else ""
    result_pass = "YES" if re.search(r"Conclusión:?\s*CUMPLE", compact, re.I) else ""

    return {
        "provider": provider,
        "manufacturer": manufacturer,
        "report_reference": report_reference,
        "report_date": report_date,
        "issued_by": "QUINEL Ltd" if "QUINEL" in full_text else "",
        "rng_report_reference": rng_report_reference,
        "rng_report_date": rng_report_date,
        "rng_issued_by": rng_issued_by,
        "bmm_revision_status": "",
        "match_with_jurisdiction_in_scope": jurisdiction,
        "general_result_is_pass": result_pass,
        "expected_games": extract_expected_count(full_text),
    }


def get_summary_text(full_text):
    """Limita la extracción de juegos a A.1.1 Resumen, antes de A.1.2 Archivos Críticos."""
    compact = re.sub(r"\s+", " ", full_text)
    summary_text = compact

    start_match = re.search(r"A\.1\.1\.?\s*Resumen", compact, re.I)
    if start_match:
        summary_text = compact[start_match.start():]

    end_match = re.search(r"A\.1\.2\.?\s*Archivos Cr[ií]ticos", summary_text, re.I)
    if end_match:
        summary_text = summary_text[:end_match.start()]

    return summary_text


def normalize_game_name(name):
    name = clean(name)
    name = re.sub(r"^ITEM\s+NOMBRE DEL\s+PRODUCTO\s+", "", name, flags=re.I).strip()
    name = re.sub(r"^ITEM\s+NOMBRE DEL PRODUCTO\s+", "", name, flags=re.I).strip()
    name = re.sub(r"\s+Código único.*$", "", name, flags=re.I).strip()
    name = re.sub(r"\s+VERSI[OÓ]N.*$", "", name, flags=re.I).strip()
    return name


def normalize_game_type(game_type):
    game_type = clean(game_type)
    game_type = re.sub(r"\s+MEDIOS SOPORTADOS.*$", "", game_type, flags=re.I).strip()
    return game_type


def extract_games_from_compact_summary(summary_text):
    """
    Extrae juegos desde tablas convertidas a texto continuo.
    Soporta tabla simple y tabla con código único.
    """
    games = []

    pattern = re.compile(
        r"\b(G\d{3})\s+"
        r"(.+?)\s+"
        r"(\d+(?:\.\d+){1,3})\s+"
        r"(.+?)\s+"
        r"HTML5",
        re.I,
    )

    for m in pattern.finditer(summary_text):
        item = clean(m.group(1)).upper()
        raw_name = normalize_game_name(m.group(2))
        version = clean(m.group(3))
        game_type = normalize_game_type(m.group(4))

        # Remueve código único si quedó pegado al nombre.
        name, unique_code = split_joined_unique_code(raw_name)

        games.append({
            "item": item,
            "game_name": name,
            "game_type": game_type or "Juegos de Línea",
            "sample": version,
            "unique_code": unique_code,
        })

    return games


def extract_games_from_lines(pages):
    """
    Fallback para PDFs donde la tabla queda partida en líneas.
    """
    games = []

    limited_pages = []
    stop = False
    for lines in pages:
        current = []
        for line in lines:
            if re.search(r"A\.1\.2|Archivos Cr[ií]ticos", line, re.I):
                stop = True
                break
            current.append(line)
        limited_pages.append(current)
        if stop:
            break

    for lines in limited_pages:
        i = 0
        while i < len(lines):
            value = clean(lines[i])

            if re.fullmatch(r"G\d{3}", value):
                item = value.upper()
                ptr = i + 1
                name_parts = []
                unique_code_parts = []

                while ptr < len(lines):
                    current = clean(lines[ptr])

                    if is_version(current):
                        break

                    if is_unique_code(current):
                        unique_code_parts.append(current)
                        ptr += 1
                        continue

                    if re.fullmatch(r"G\d{3}", current):
                        break

                    if re.search(r"ITEM|VERSI[OÓ]N|TIPO DE|MEDIOS SOPORTADOS|Código único", current, re.I):
                        ptr += 1
                        continue

                    # Si luce como parte de código único partido, lo guardamos como código.
                    if re.fullmatch(r"[A-Za-z0-9]{8,}", current) or re.fullmatch(r"[A-Za-z0-9]{6,}_[0-9.]+", current):
                        unique_code_parts.append(current)
                    else:
                        name_parts.append(current)

                    ptr += 1

                if ptr < len(lines) and is_version(lines[ptr]):
                    version = clean(lines[ptr])
                    ptr += 1

                    game_type_parts = []
                    while ptr < len(lines):
                        current = clean(lines[ptr])
                        if re.fullmatch(r"G\d{3}", current):
                            break
                        if re.search(r"HTML5|MEDIOS|SOPORTADOS|A\.1\.2", current, re.I):
                            break
                        game_type_parts.append(current)
                        ptr += 1

                    game_type = normalize_game_type(" ".join(game_type_parts))
                    name = normalize_game_name(" ".join(name_parts))
                    unique_code = "".join(unique_code_parts).replace(" ", "")

                    if name and version:
                        games.append({
                            "item": item,
                            "game_name": name,
                            "game_type": game_type or "Juegos de Línea",
                            "sample": version,
                            "unique_code": unique_code,
                        })

                    i = ptr
                    continue

            i += 1

    return games


def dedupe_games(games):
    """
    Deduplica por item. Si el mismo item sale por dos métodos,
    se queda con el registro más completo.
    """
    by_item = {}

    for game in games:
        item = game.get("item", "")
        if not item:
            continue

        current = by_item.get(item)
        if current is None:
            by_item[item] = game
            continue

        current_score = sum(bool(current.get(k)) for k in ["game_name", "game_type", "sample", "unique_code"])
        new_score = sum(bool(game.get(k)) for k in ["game_name", "game_type", "sample", "unique_code"])

        if new_score > current_score:
            by_item[item] = game

    return [by_item[item] for item in sorted(by_item.keys())]


def extract_games(full_text, pages):
    summary_text = get_summary_text(full_text)

    games = []
    games.extend(extract_games_from_compact_summary(summary_text))
    games.extend(extract_games_from_lines(pages))

    cleaned_games = []
    for game in dedupe_games(games):
        name = clean(game.get("game_name"))
        sample = clean(game.get("sample"))
        item = clean(game.get("item"))

        if not re.fullmatch(r"G\d{3}", item):
            continue

        if not name or not sample:
            continue

        # Evita falsos positivos por texto excesivamente largo.
        if len(name) > 120:
            logging.warning("Juego descartado por nombre demasiado largo: %s", name)
            continue

        cleaned_games.append({
            "game_name": name,
            "game_type": clean(game.get("game_type")) or "Juegos de Línea",
            "sample": sample,
            "unique_code": clean(game.get("unique_code")),
        })

    return cleaned_games


def merged_value(ws, row, col):
    cell = ws.cell(row, col)
    if cell.value is not None:
        return cell.value

    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return ws.cell(merged.min_row, merged.min_col).value

    return None


def find_sheet(wb):
    if "B2B table" in wb.sheetnames:
        return wb["B2B table"]

    for ws in wb.worksheets:
        for r in range(1, min(ws.max_row, 30) + 1):
            values = [clean(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
            if "Game Provider" in values and "Game Name" in values:
                return ws

    raise ValueError("No encontré la hoja B2B table ni una hoja con los encabezados esperados.")


def find_header_row(ws):
    for r in range(1, min(ws.max_row, 40) + 1):
        values = [clean(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if "Game Provider" in values and "Game Name" in values and "Report Reference" in values:
            return r

    raise ValueError("No encontré la fila de encabezados.")


def copy_row_format(ws, source_row, target_row):
    for col in range(1, ws.max_column + 1):
        source_cell = ws.cell(source_row, col)
        target_cell = ws.cell(target_row, col)

        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)

        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)

    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def fill_excel(template_path, rows, output_path):
    template_path = Path(template_path)
    output_path = Path(output_path)

    if not template_path.exists():
        raise FileNotFoundError(f"No existe la plantilla: {template_path}")

    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(template_path)
    ws = find_sheet(wb)
    header_row = find_header_row(ws)
    marker_row = header_row - 1
    data_row = header_row + 1

    fillable_columns = {}

    for col in range(1, ws.max_column + 1):
        header = clean(ws.cell(header_row, col).value)
        marker = clean(merged_value(ws, marker_row, col)).lower()

        if marker == CLIENT_MARKER or header == "Sample":
            fillable_columns[col] = header

    if not fillable_columns:
        raise ValueError("No encontré columnas editables en la plantilla.")

    style_source_row = data_row

    # Limpieza robusta:
    # La plantilla puede venir con juegos residuales de procesamientos anteriores.
    # Para evitar que el Excel final conserve filas viejas, se deja una sola fila base
    # para copiar formato y se eliminan todas las filas de datos existentes debajo.
    if ws.max_row > data_row:
        ws.delete_rows(data_row + 1, ws.max_row - data_row)

    # Limpia toda la fila base, no solo columnas editables, para evitar residuos.
    for col in range(1, ws.max_column + 1):
        ws.cell(data_row, col).value = None

    rows_to_create = max(len(rows), 1)

    for idx in range(rows_to_create):
        excel_row = data_row + idx
        copy_row_format(ws, style_source_row, excel_row)

        # Limpia la fila completa antes de escribir.
        for col in range(1, ws.max_column + 1):
            ws.cell(excel_row, col).value = None

    for idx, row in enumerate(rows):
        excel_row = data_row + idx

        for col, header in fillable_columns.items():
            if header in ALWAYS_BLANK:
                ws.cell(excel_row, col).value = None
                continue

            field = FIELD_MAP.get(header)
            if not field:
                continue

            value = row.get(field, "")
            ws.cell(excel_row, col).value = value or None

    wb.save(output_path)


def build_rows_for_pdf(pdf):
    pdf_path = Path(pdf)
    full_text, pages = read_pdf_text(pdf_path)
    header = extract_header(full_text, pages)
    games = extract_games(full_text, pages)

    rows = []
    for game in games:
        row = {}
        row.update(header)
        row.update(game)
        rows.append(row)

    expected = header.get("expected_games")
    return rows, {
        "pdf": pdf_path.name,
        "report_reference": header.get("report_reference", ""),
        "expected_games": expected if expected is not None else "",
        "extracted_games": len(games),
        "status": "OK" if expected is None or expected == len(games) else "REVISAR",
        "message": "" if expected is None or expected == len(games) else f"Esperados {expected}, extraídos {len(games)}",
    }


def write_audit_csv(audit_rows, audit_path):
    audit_path = Path(audit_path)
    audit_path.parent.mkdir(parents=True, exist_ok=True)

    columns = ["pdf", "report_reference", "expected_games", "extracted_games", "status", "message"]

    with audit_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=columns)
        writer.writeheader()
        for row in audit_rows:
            writer.writerow({col: row.get(col, "") for col in columns})


def resolve_pdfs(pdf_args, pdf_dir):
    pdfs = []

    for pdf in pdf_args or []:
        pdfs.append(Path(pdf))

    if pdf_dir:
        pdf_dir = Path(pdf_dir)
        if not pdf_dir.exists():
            raise FileNotFoundError(f"No existe la carpeta de PDFs: {pdf_dir}")
        pdfs.extend(sorted(pdf_dir.glob("*.pdf")))

    # Deduplicar preservando orden.
    seen = set()
    unique_pdfs = []
    for pdf in pdfs:
        key = str(pdf.resolve())
        if key not in seen:
            unique_pdfs.append(pdf)
            seen.add(key)

    if not unique_pdfs:
        raise ValueError("No se recibió ningún PDF. Usa --pdf o --pdf-dir.")

    return unique_pdfs


def process(template, pdfs, output, audit=None, strict=False):
    all_rows = []
    audit_rows = []

    for pdf in pdfs:
        try:
            rows, audit_row = build_rows_for_pdf(pdf)
            all_rows.extend(rows)
            audit_rows.append(audit_row)

            logging.info(
                "%s: %s juegos extraídos. Estado: %s",
                audit_row["pdf"],
                audit_row["extracted_games"],
                audit_row["status"],
            )

            if strict and audit_row["status"] != "OK":
                raise ValueError(audit_row["message"])

        except Exception as exc:
            audit_row = {
                "pdf": Path(pdf).name,
                "report_reference": "",
                "expected_games": "",
                "extracted_games": 0,
                "status": "ERROR",
                "message": str(exc),
            }
            audit_rows.append(audit_row)
            logging.exception("Error procesando %s", pdf)

            if strict:
                raise

    fill_excel(template, all_rows, output)

    if audit:
        write_audit_csv(audit_rows, audit)

    logging.info("Excel generado: %s", output)
    logging.info("Total de filas: %s", len(all_rows))

    revisar = [row for row in audit_rows if row["status"] != "OK"]
    if revisar:
        logging.warning("Hay %s PDF(s) con estado REVISAR/ERROR. Consulta el audit CSV.", len(revisar))

    return audit_rows


def parse_args():
    parser = argparse.ArgumentParser(
        description="Genera Excel B2B de integración de juegos a partir de certificados PDF para Perú."
    )
    parser.add_argument("--template", required=True, help="Ruta de la plantilla Excel.")
    parser.add_argument("--pdf", action="append", help="Ruta de un PDF. Se puede repetir.")
    parser.add_argument("--pdf-dir", help="Carpeta con PDFs. Procesa todos los *.pdf.")
    parser.add_argument("--output", required=True, help="Ruta del Excel generado.")
    parser.add_argument("--audit", help="Ruta opcional para CSV de auditoría.")
    parser.add_argument(
        "--strict",
        action="store_true",
        help="Detiene el proceso si un PDF falla o si el conteo esperado no coincide.",
    )
    parser.add_argument(
        "--log-level",
        default="INFO",
        choices=["DEBUG", "INFO", "WARNING", "ERROR"],
        help="Nivel de log.",
    )
    return parser.parse_args()


if __name__ == "__main__":
    args = parse_args()

    logging.basicConfig(
        level=getattr(logging, args.log_level),
        format="%(levelname)s | %(message)s",
    )

    pdfs = resolve_pdfs(args.pdf, args.pdf_dir)

    process(
        template=args.template,
        pdfs=pdfs,
        output=args.output,
        audit=args.audit,
        strict=args.strict,
    )
