from pathlib import Path
import re
import tempfile
import shutil

import fitz
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from generar_excel import process


# =========================
# CONFIGURACIÓN GENERAL
# =========================

APP_TITLE = "Certificador de Juegos - Perú"
LOGO_PATH = Path("assets/logo_micasino.png")
DEFAULT_TEMPLATE_PATH = Path("B2B TEMPLATE- GAMES INTEGRATIONS.xlsx")

MI_CASINO_YELLOW = "#FFC629"
MI_CASINO_BLACK = "#0B0B0B"
MI_CASINO_WHITE = "#FFFFFF"
MI_CASINO_BG = "#FFF9EA"
MI_CASINO_SOFT = "#FFF3C4"


# =========================
# UTILIDADES GENERALES
# =========================


def clean(value):
    """Normaliza espacios y convierte None a texto vacío."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


@st.cache_data(show_spinner=False)
def get_logo_bytes():
    if LOGO_PATH.exists():
        return LOGO_PATH.read_bytes()
    return None


def find_template():
    """Busca la plantilla esperada. Si no existe, intenta encontrar una .xlsx similar."""
    if DEFAULT_TEMPLATE_PATH.exists():
        return DEFAULT_TEMPLATE_PATH

    candidates = sorted(Path(".").glob("*.xlsx"))
    for candidate in candidates:
        if "B2B" in candidate.name.upper() and "GAME" in candidate.name.upper():
            return candidate

    return DEFAULT_TEMPLATE_PATH


def load_file_bytes(path):
    with open(path, "rb") as file:
        return file.read()


# =========================
# LÓGICA MINCETUR
# =========================


def read_pdf_text(pdf_path):
    """Lee texto de un PDF usando PyMuDF."""
    pdf_path = Path(pdf_path)

    if not pdf_path.exists():
        raise FileNotFoundError(f"No existe el PDF: {pdf_path}")

    pages = []

    with fitz.open(str(pdf_path)) as doc:
        if doc.is_encrypted:
            raise ValueError(f"El PDF está cifrado/protegido: {pdf_path.name}")

        for page in doc:
            lines = [clean(line) for line in page.get_text("text").splitlines() if clean(line)]
            pages.append(lines)

    full_text = "\n".join("\n".join(lines) for lines in pages)

    if not clean(full_text):
        raise ValueError(f"No se pudo extraer texto del PDF: {pdf_path.name}")

    return full_text, pages


def detect_tipo_componente(full_text):
    """Detecta el tipo de componente mencionado por la Resolución Directoral."""
    text = clean(full_text).upper()

    if "PROGRAMAS DE JUEGO" in text or "PROGRAMA DE JUEGO" in text:
        return "Programa de juego"

    if "PLATAFORMA TECNOLÓGICA" in text or "PLATAFORMA TECNOLOGICA" in text:
        return "Plataforma tecnológica"

    if "SISTEMA PROGRESIVO" in text:
        return "Sistema progresivo"

    if "CASINO EN VIVO" in text:
        return "Modalidad de juego de casino en vivo"

    return "No identificado"


def normalize_broken_code(value):
    """
    Une códigos que el PDF puede partir en varias líneas.

    Ejemplo:
    vs20payanyvol_cv1 13 -> vs20payanyvol_cv113
    """
    value = clean(value)
    value = re.sub(r"_cv\s+(\d+)", r"_cv\1", value, flags=re.I)
    value = re.sub(r"_cv(\d+)\s+(\d+)", r"_cv\1\2", value, flags=re.I)
    return value


def extract_resolution_manufacturer(full_text):
    """
    Obtiene el fabricante/proveedor principal de la Resolución Directoral.

    Esto evita que el nombre del fabricante se pegue al nombre comercial cuando
    PyMuPDF extrae la tabla como texto continuo.
    """
    compact = re.sub(r"\s+", " ", full_text)

    patterns = [
        r"presentado\s+por\s+la\s+empresa\s+(.+?)\s+en\s+el\s+que\s+solicita",
        r"a\s+solicitud\s+del\s+proveedor\s+de\s+servicios\s+vinculados\s+(.+?),\s+la\s+inscripci[oó]n",
        r"a\s+solicitud\s+de\s+(.+?),\s+la\s+inscripci[oó]n",
    ]

    for pattern in patterns:
        match = re.search(pattern, compact, re.I)
        if match:
            manufacturer = clean(match.group(1)).strip(" ,.;:")
            if manufacturer:
                return manufacturer

    return ""


def strip_known_manufacturer(value, manufacturer):
    """
    Quita el fabricante del inicio de una cadena y deja solo el nombre comercial.

    Ejemplo:
    NERINE SERVICES LIMITED Limbo+ -> Limbo+
    SERVICES LIMITED Limbo+ -> Limbo+
    """
    value = clean(value)
    manufacturer = clean(manufacturer)

    if not value:
        return ""

    if manufacturer:
        manufacturer_pattern = r"\s+".join(re.escape(token) for token in manufacturer.split())
        value = re.sub(rf"^{manufacturer_pattern}\s+", "", value, flags=re.I).strip()

        # Si por el corte de columnas quedó solo la cola del fabricante,
        # por ejemplo "SERVICES LIMITED Limbo+", se elimina esa cola también.
        tokens = manufacturer.split()
        for start in range(1, len(tokens)):
            suffix = " ".join(tokens[start:])
            suffix_pattern = r"\s+".join(re.escape(token) for token in suffix.split())
            new_value = re.sub(rf"^{suffix_pattern}\s+", "", value, flags=re.I).strip()
            if new_value != value:
                value = new_value
                break

    # Fallback genérico para fabricantes que terminan en LIMITED, LTD, LLC, INC, etc.
    value = re.sub(
        r"^(?:[A-ZÁÉÍÓÚÑ0-9&.,'’\-]+\s+){0,8}"
        r"(?:LIMITED|LTD\.?|LLC|INC\.?|CORP\.?|S\.?A\.?C?\.?|GMBH)\s+",
        "",
        value,
        flags=re.I,
    ).strip()

    return value


def extract_mincetur_resolution_rows(pdf_path):
    """
    Extrae registros homologados desde Resoluciones Directorales MINCETUR.

    Tabla esperada en la parte resolutiva:
    N° REGISTRO | NOMBRE DEL FABRICANTE | NOMBRE COMERCIAL DEL JUEGO | VERSIÓN | CÓDIGO DE IDENTIFICACIÓN
    """
    full_text, _ = read_pdf_text(pdf_path)
    tipo_componente = detect_tipo_componente(full_text)
    manufacturer = extract_resolution_manufacturer(full_text)

    compact = re.sub(r"\s+", " ", full_text)

    article_match = re.search(r"Artículo\s+1\.?-", compact, re.I)
    if article_match:
        search_text = compact[article_match.start():]
    else:
        search_text = compact

    rows = []

    # En el texto extraído, las columnas quedan unidas. Por eso se toma todo
    # entre N° REGISTRO y VERSIÓN como "middle" y luego se elimina el fabricante.
    pattern = re.compile(
        r"\b(?P<n>\d+)\s+"
        r"(?P<registro>PJ\d{7})\s+"
        r"(?P<middle>.+?)\s+"
        r"(?P<version>(?:cv|v)?\d+(?:\.\d+)+)\s+"
        r"(?P<codigo>[A-Za-z0-9]+(?:[_-][A-Za-z0-9]+)*(?:\s+\d+)?)(?=\s+\d+\s+PJ\d{7}|\s+Artículo\s+2|$)",
        re.I,
    )

    seen = set()

    for match in pattern.finditer(search_text):
        middle = clean(match.group("middle"))
        nombre = strip_known_manufacturer(middle, manufacturer)
        codigo = normalize_broken_code(match.group("codigo"))
        registro = clean(match.group("registro"))

        nombre = re.sub(r"^NOMBRE\s+COMERCIAL\s+DEL\s+JUEGO\s+", "", nombre, flags=re.I).strip()
        nombre = re.sub(r"^NOMBRE\s+COMERCIAL\s+", "", nombre, flags=re.I).strip()

        if not nombre or len(nombre) > 100:
            continue

        key = (registro, codigo, nombre.lower())
        if key in seen:
            continue
        seen.add(key)

        rows.append(
            {
                "tipo_componente": tipo_componente,
                "nombre_comercial": nombre,
                "codigo_identificacion_fabricante": codigo,
                "registro_mincetur": registro,
                "pdf": Path(pdf_path).name,
            }
        )

    return rows


def write_mincetur_excel(rows, output_path):
    """Escribe Excel con el formato solicitado para Resoluciones Directorales."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Resoluciones MINCETUR"

    headers = [
        "N° correlativo",
        "Tipo de componente",
        "Nombre comercial",
        "Código de identificación (fabricante)",
        "Número de registro del modelo otorgado por MINCETUR",
    ]

    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="D9EAD3")
    header_font = Font(bold=True, color="000000")
    border = Border(
        left=Side(style="thin", color="B7B7B7"),
        right=Side(style="thin", color="B7B7B7"),
        top=Side(style="thin", color="B7B7B7"),
        bottom=Side(style="thin", color="B7B7B7"),
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for idx, row in enumerate(rows, start=1):
        ws.append(
            [
                idx,
                row.get("tipo_componente", ""),
                row.get("nombre_comercial", ""),
                row.get("codigo_identificacion_fabricante", ""),
                row.get("registro_mincetur", ""),
            ]
        )

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["D"].width = 42
    ws.column_dimensions["E"].width = 58

    wb.save(output_path)


def process_mincetur_resolutions(pdfs, output_path):
    """Procesa una o varias Resoluciones Directorales y genera el Excel regulatorio."""
    all_rows = []
    audit_rows = []

    for pdf in pdfs:
        try:
            rows = extract_mincetur_resolution_rows(pdf)
            all_rows.extend(rows)

            audit_rows.append(
                {
                    "pdf": Path(pdf).name,
                    "registros_extraidos": len(rows),
                    "status": "OK" if rows else "REVISAR",
                    "message": "" if rows else "No se encontraron registros MINCETUR.",
                }
            )

        except Exception as exc:
            audit_rows.append(
                {
                    "pdf": Path(pdf).name,
                    "registros_extraidos": 0,
                    "status": "ERROR",
                    "message": str(exc),
                }
            )

    write_mincetur_excel(all_rows, output_path)
    return all_rows, audit_rows


# =========================
# ESTADO DE SESIÓN
# =========================


def init_session_state():
    defaults = {
        # Certificados GLI / QUINEL.
        "processed": False,
        "excel_bytes": None,
        "audit_bytes": None,
        "audit_df": None,
        "audit_rows": [],
        "excel_downloaded": False,
        "audit_downloaded": False,
        "uploader_key": 0,
        "last_error": None,
        # Resoluciones MINCETUR.
        "mincetur_processed": False,
        "mincetur_excel_bytes": None,
        "mincetur_rows": [],
        "mincetur_audit_rows": [],
        "mincetur_downloaded": False,
        "mincetur_uploader_key": 0,
    }

    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


def reset_certificate_results():
    st.session_state.processed = False
    st.session_state.excel_bytes = None
    st.session_state.audit_bytes = None
    st.session_state.audit_df = None
    st.session_state.audit_rows = []
    st.session_state.excel_downloaded = False
    st.session_state.audit_downloaded = False
    st.session_state.last_error = None


def reset_mincetur_results():
    st.session_state.mincetur_processed = False
    st.session_state.mincetur_excel_bytes = None
    st.session_state.mincetur_rows = []
    st.session_state.mincetur_audit_rows = []
    st.session_state.mincetur_downloaded = False
    st.session_state.mincetur_uploader_key += 1


def reset_all_results():
    reset_certificate_results()
    reset_mincetur_results()
    st.session_state.uploader_key += 1


def mark_excel_downloaded():
    st.session_state.excel_downloaded = True


def mark_audit_downloaded():
    st.session_state.audit_downloaded = True


def mark_mincetur_downloaded():
    st.session_state.mincetur_downloaded = True


# =========================
# UI / ESTILOS
# =========================


def render_css():
    st.markdown(
        f"""
        <style>
            :root {{
                --mc-yellow: {MI_CASINO_YELLOW};
                --mc-black: {MI_CASINO_BLACK};
                --mc-white: {MI_CASINO_WHITE};
                --mc-bg: {MI_CASINO_BG};
                --mc-soft: {MI_CASINO_SOFT};
            }}

            .stApp {{
                background: linear-gradient(180deg, #FFFFFF 0%, #FFF9EA 100%);
            }}

            [data-testid="stSidebar"] {{
                background: #0B0B0B;
                color: white;
            }}

            [data-testid="stSidebar"] * {{
                color: white;
            }}

            .brand-hero {{
                background: linear-gradient(135deg, #0B0B0B 0%, #1E1E1E 60%, #FFC629 100%);
                padding: 28px 32px;
                border-radius: 26px;
                border: 2px solid #FFC629;
                box-shadow: 0 18px 45px rgba(0,0,0,0.16);
                margin-bottom: 22px;
            }}

            .brand-title {{
                color: white;
                font-size: 2.25rem;
                font-weight: 900;
                margin: 0;
                line-height: 1.1;
            }}

            .brand-subtitle {{
                color: #F4F4F4;
                font-size: 1rem;
                margin-top: 8px;
                max-width: 900px;
            }}

            .status-card {{
                background: white;
                border: 1px solid #E6E6E6;
                border-left: 7px solid #FFC629;
                border-radius: 18px;
                padding: 18px 20px;
                margin: 14px 0;
                box-shadow: 0 10px 25px rgba(0,0,0,0.05);
            }}

            .success-card {{
                background: #ECFDF3;
                border: 1px solid #ABEFC6;
                border-left: 7px solid #12B76A;
                color: #067647;
                border-radius: 18px;
                padding: 18px 20px;
                margin: 14px 0;
            }}

            .warning-card {{
                background: #FFFAEB;
                border: 1px solid #FEDF89;
                border-left: 7px solid #F79009;
                color: #93370D;
                border-radius: 18px;
                padding: 18px 20px;
                margin: 14px 0;
            }}

            .error-card {{
                background: #FEF3F2;
                border: 1px solid #FECDCA;
                border-left: 7px solid #F04438;
                color: #B42318;
                border-radius: 18px;
                padding: 18px 20px;
                margin: 14px 0;
            }}

            .step-row {{
                background: white;
                border-radius: 999px;
                border: 1px solid #E6E6E6;
                padding: 10px 16px;
                margin: 8px 0 24px 0;
                box-shadow: 0 8px 20px rgba(0,0,0,0.04);
                font-weight: 700;
            }}

            .mc-pill {{
                display: inline-block;
                background: #FFC629;
                color: #0B0B0B;
                padding: 6px 12px;
                border-radius: 999px;
                font-weight: 800;
                margin-right: 8px;
            }}

            div.stButton > button,
            div.stDownloadButton > button {{
                border-radius: 14px;
                min-height: 3rem;
                font-weight: 800;
                border: 1px solid #0B0B0B;
            }}

            div.stButton > button[kind="primary"],
            div.stDownloadButton > button[kind="primary"] {{
                background: #FFC629;
                color: #0B0B0B;
                border: 1px solid #0B0B0B;
            }}

            div.stButton > button:disabled,
            div.stDownloadButton > button:disabled {{
                background: #D9D9D9 !important;
                color: #555555 !important;
                border: 1px solid #999999 !important;
                opacity: 1 !important;
            }}

            .sidebar-action-note {{
                font-size: 0.85rem;
                color: #D1D5DB !important;
            }}
                        /* ==============================
               Forzar visual light mode
               ============================== */

            html,
            body,
            .stApp,
            [data-testid="stAppViewContainer"],
            [data-testid="stHeader"] {{
                background-color: #FFF9EA !important;
                color: #111827 !important;
            }}

            .block-container {{
                background-color: transparent !important;
                color: #111827 !important;
            }}

            h1, h2, h3, h4, h5, h6,
            p, span, label, div {{
                color: inherit;
            }}

            /* Textos principales fuera del sidebar */
            [data-testid="stAppViewContainer"] h1,
            [data-testid="stAppViewContainer"] h2,
            [data-testid="stAppViewContainer"] h3,
            [data-testid="stAppViewContainer"] p,
            [data-testid="stAppViewContainer"] label,
            [data-testid="stAppViewContainer"] span {{
                color: #111827 !important;
            }}

            /* File uploader */
            [data-testid="stFileUploader"] {{
                background-color: #F3F4F6 !important;
                border-radius: 12px !important;
                color: #111827 !important;
            }}

            [data-testid="stFileUploader"] section {{
                background-color: #F3F4F6 !important;
                border: 1px solid #E5E7EB !important;
                color: #111827 !important;
            }}

            [data-testid="stFileUploader"] button {{
                background-color: #FFFFFF !important;
                color: #111827 !important;
                border: 1px solid #D1D5DB !important;
            }}

            [data-testid="stFileUploader"] small,
            [data-testid="stFileUploader"] span,
            [data-testid="stFileUploader"] p {{
                color: #6B7280 !important;
            }}

            /* Inputs, selects y áreas de texto */
            input,
            textarea,
            select {{
                background-color: #FFFFFF !important;
                color: #111827 !important;
                border-color: #D1D5DB !important;
            }}

            /* Métricas */
            [data-testid="stMetric"] {{
                background-color: transparent !important;
                color: #111827 !important;
            }}

            [data-testid="stMetric"] label,
            [data-testid="stMetric"] div {{
                color: #111827 !important;
            }}

            /* Dataframes / tablas */
            [data-testid="stDataFrame"],
            [data-testid="stTable"] {{
                background-color: #FFFFFF !important;
                color: #111827 !important;
                border-radius: 12px !important;
            }}

            [data-testid="stDataFrame"] * {{
                color: #111827 !important;
            }}

            /* Expanders */
            [data-testid="stExpander"] {{
                background-color: #FFFFFF !important;
                color: #111827 !important;
                border: 1px solid #E5E7EB !important;
                border-radius: 12px !important;
            }}

            [data-testid="stExpander"] * {{
                color: #111827 !important;
            }}

            /* Tabs */
            button[data-baseweb="tab"] {{
                background-color: #FFFFFF !important;
                color: #111827 !important;
                border-radius: 999px !important;
            }}

            button[data-baseweb="tab"][aria-selected="true"] {{
                background-color: var(--mc-yellow) !important;
                color: #000000 !important;
                font-weight: 800 !important;
            }}

            button[data-baseweb="tab"] p {{
                color: inherit !important;
            }}

            /* Botones principales */
            div.stButton > button,
            div.stDownloadButton > button {{
                background-color: var(--mc-yellow) !important;
                color: #000000 !important;
                border: 2px solid var(--mc-yellow) !important;
                border-radius: 14px !important;
                font-weight: 800 !important;
            }}

            div.stButton > button p,
            div.stDownloadButton > button p {{
                color: #000000 !important;
                font-weight: 800 !important;
            }}

            div.stButton > button:disabled,
            div.stDownloadButton > button:disabled {{
                background-color: #FFE08A !important;
                color: #000000 !important;
                border-color: #FFE08A !important;
                opacity: 0.75 !important;
            }}

            div.stButton > button:disabled p,
            div.stDownloadButton > button:disabled p {{
                color: #000000 !important;
            }}

            /* Sidebar se mantiene oscuro */
            [data-testid="stSidebar"] {{
                background: #0B0B0B !important;
                color: #FFFFFF !important;
            }}

            [data-testid="stSidebar"] h1,
            [data-testid="stSidebar"] h2,
            [data-testid="stSidebar"] h3,
            [data-testid="stSidebar"] p,
            [data-testid="stSidebar"] span,
            [data-testid="stSidebar"] label,
            [data-testid="stSidebar"] div {{
                color: #FFFFFF !important;
            }}

            /* Botón limpiar resultados en sidebar */
            section[data-testid="stSidebar"] div.stButton > button {{
                background-color: var(--mc-yellow) !important;
                color: #000000 !important;
                border: 2px solid var(--mc-yellow) !important;
                border-radius: 14px !important;
                font-weight: 800 !important;
                opacity: 1 !important;
            }}

            section[data-testid="stSidebar"] div.stButton > button p {{
                color: #000000 !important;
                font-weight: 800 !important;
            }}
.hero-card h1,
.hero-card .hero-title {{
    color: #FFFFFF !important;
    text-shadow: 0 2px 8px rgba(0, 0, 0, 0.55);
}}

.hero-card p,
.hero-card .hero-subtitle {{
    color: #FFFFFF !important;
    font-weight: 600 !important;
}}
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_header():
    logo_bytes = get_logo_bytes()

    with st.container():
        if logo_bytes:
            col_logo, col_text = st.columns([1, 6])
            with col_logo:
                st.image(logo_bytes, width=90)
            with col_text:
                st.markdown(
                    f"""
                    <div class="brand-hero">
                        <h1 class="brand-title">{APP_TITLE}</h1>
                        <div class="brand-subtitle">
                            Carga certificados de juegos y Resoluciones Directorales MINCETUR.
                            Genera archivos Excel listos para validación y control regulatorio.
                        </div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )
        else:
            st.markdown(
                f"""
                <div class="brand-hero">
                    <h1 class="brand-title">{APP_TITLE}</h1>
                    <div class="brand-subtitle">
                        Carga certificados de juegos y Resoluciones Directorales MINCETUR.
                        Genera archivos Excel listos para validación y control regulatorio.
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )


def render_sidebar():
    with st.sidebar:
        logo_bytes = get_logo_bytes()
        if logo_bytes:
            st.image(logo_bytes, width=110)

        st.markdown("## MiCasino.com")
        st.caption("Herramienta interna de certificación y homologación Perú")

        st.divider()

        template_path = find_template()
        st.markdown("### Plantilla B2B")
        if template_path.exists():
            st.success(template_path.name)
        else:
            st.error("No se encontró la plantilla Excel B2B.")

        st.divider()

        st.markdown("### Entidades certificadoras soportadas")
        st.markdown("🟢 QUINEL Ltd")
        st.markdown("🟢 Gaming Laboratories International (GLI)")

        st.divider()

        st.markdown("### Documento regulatorio soportado")
        st.markdown("🟢 Resolución Directoral MINCETUR")

        st.divider()

        st.markdown("### Acciones")
        if st.button("Limpiar resultados", use_container_width=True):
            reset_all_results()
            st.rerun()

        st.markdown(
            '<p class="sidebar-action-note">Limpia resultados, descargas y archivos cargados visualmente.</p>',
            unsafe_allow_html=True,
        )


def style_audit_table(df):
    if df.empty or "status" not in df.columns:
        return df

    def color_status(row):
        status = str(row.get("status", "")).upper()
        if status == "OK":
            return ["background-color: #ECFDF3; color: #067647"] * len(row)
        if status == "REVISAR":
            return ["background-color: #FFFAEB; color: #93370D"] * len(row)
        if status == "ERROR":
            return ["background-color: #FEF3F2; color: #B42318"] * len(row)
        return [""] * len(row)

    return df.style.apply(color_status, axis=1)


# =========================
# TABS CERTIFICADOS
# =========================


def render_certificates_tab():
    st.markdown(
        """
        <div class="step-row">
            <span class="mc-pill">1</span>Cargar PDFs
            <span class="mc-pill">2</span>Procesar certificados
            <span class="mc-pill">3</span>Validar auditoría
            <span class="mc-pill">4</span>Descargar archivos
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.subheader("Carga y procesamiento de certificados")
    st.write("Sube uno o varios certificados GLI o QUINEL. La herramienta generará el Excel B2B y un CSV de auditoría.")

    template_path = find_template()

    uploaded_files = st.file_uploader(
        "Arrastra o selecciona tus certificados PDF",
        type=["pdf"],
        accept_multiple_files=True,
        key=f"pdf_uploader_{st.session_state.uploader_key}",
    )

    if not uploaded_files and not st.session_state.processed:
        st.markdown(
            """
            <div class="status-card">
                Carga certificados PDF para iniciar el proceso. Después podrás revisar la auditoría
                y descargar el Excel completado junto con el CSV de control.
            </div>
            """,
            unsafe_allow_html=True,
        )

    if uploaded_files:
        st.markdown(f"**PDFs cargados:** {len(uploaded_files)}")
        with st.expander("Ver archivos cargados", expanded=True):
            for uploaded_file in uploaded_files:
                st.write(f"📄 {uploaded_file.name}")

        process_clicked = st.button(
            "Procesar certificados",
            type="primary",
            use_container_width=True,
            disabled=not template_path.exists(),
        )

        if process_clicked:
            reset_certificate_results()

            with st.spinner("Procesando certificados..."):
                try:
                    with tempfile.TemporaryDirectory() as tmpdir:
                        tmpdir = Path(tmpdir)

                        pdf_dir = tmpdir / "pdfs"
                        pdf_dir.mkdir(parents=True, exist_ok=True)

                        pdf_paths = []
                        for uploaded_file in uploaded_files:
                            pdf_path = pdf_dir / uploaded_file.name
                            with open(pdf_path, "wb") as file:
                                file.write(uploaded_file.getbuffer())
                            pdf_paths.append(pdf_path)

                        template_copy = tmpdir / template_path.name
                        shutil.copy(template_path, template_copy)

                        output_excel = tmpdir / "B2B_TEMPLATE_GAMES_INTEGRATIONS_COMPLETADO.xlsx"
                        output_audit = tmpdir / "auditoria_certificados.csv"

                        audit_rows = process(
                            template=template_copy,
                            pdfs=pdf_paths,
                            output=output_excel,
                            audit=output_audit,
                            strict=False,
                        )

                        st.session_state.excel_bytes = load_file_bytes(output_excel)
                        st.session_state.audit_bytes = load_file_bytes(output_audit)
                        st.session_state.audit_rows = audit_rows
                        st.session_state.audit_df = pd.DataFrame(audit_rows)
                        st.session_state.processed = True
                        st.session_state.excel_downloaded = False
                        st.session_state.audit_downloaded = False
                        st.session_state.last_error = None
                        st.session_state.uploader_key += 1

                    st.success("Certificados procesados correctamente.")
                    st.rerun()

                except Exception as exc:
                    st.session_state.last_error = str(exc)
                    st.error(f"Ocurrió un error procesando los certificados: {exc}")

    if st.session_state.last_error:
        st.markdown(
            f"""
            <div class="error-card">
                <strong>Error:</strong> {st.session_state.last_error}
            </div>
            """,
            unsafe_allow_html=True,
        )

    if st.session_state.processed:
        render_certificates_result_summary(show_downloads=True)


def render_certificates_result_summary(show_downloads=False):
    df = st.session_state.audit_df
    if df is None:
        df = pd.DataFrame(st.session_state.audit_rows)

    if df.empty:
        st.warning("No hay datos de auditoría disponibles.")
        return

    total_pdfs = len(df)
    ok_count = int((df["status"] == "OK").sum()) if "status" in df.columns else 0
    revisar_count = int((df["status"] == "REVISAR").sum()) if "status" in df.columns else 0
    error_count = int((df["status"] == "ERROR").sum()) if "status" in df.columns else 0
    total_games = int(pd.to_numeric(df.get("extracted_games", 0), errors="coerce").fillna(0).sum())

    st.divider()
    st.markdown("### Resultado listo")

    col1, col2, col3, col4, col5 = st.columns(5)
    col1.metric("PDFs", total_pdfs)
    col2.metric("OK", ok_count)
    col3.metric("Revisar", revisar_count)
    col4.metric("Errores", error_count)
    col5.metric("Juegos", total_games)

    if error_count > 0:
        st.markdown(
            """
            <div class="error-card">
                Hay certificados con error. Revisa la auditoría antes de usar el Excel.
            </div>
            """,
            unsafe_allow_html=True,
        )
    elif revisar_count > 0:
        st.markdown(
            """
            <div class="warning-card">
                Hay certificados marcados como REVISAR. El Excel fue generado, pero conviene validar esos casos manualmente.
            </div>
            """,
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            """
            <div class="success-card">
                Todos los certificados fueron procesados correctamente. Puedes descargar el Excel y el CSV de auditoría.
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.markdown("### Vista rápida de auditoría")
    st.dataframe(style_audit_table(df), use_container_width=True, hide_index=True)

    if show_downloads:
        st.markdown("### Descargas")
        col_excel, col_audit = st.columns(2)

        with col_excel:
            st.download_button(
                label="Descargar Excel B2B completado",
                data=st.session_state.excel_bytes,
                file_name="B2B_TEMPLATE_GAMES_INTEGRATIONS_COMPLETADO.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
                disabled=st.session_state.excel_downloaded,
                on_click=mark_excel_downloaded,
            )
            if st.session_state.excel_downloaded:
                st.caption("Excel descargado. Botón bloqueado para evitar descargas duplicadas.")

        with col_audit:
            st.download_button(
                label="Descargar CSV de auditoría",
                data=st.session_state.audit_bytes,
                file_name="auditoria_certificados.csv",
                mime="text/csv",
                use_container_width=True,
                disabled=st.session_state.audit_downloaded,
                on_click=mark_audit_downloaded,
            )
            if st.session_state.audit_downloaded:
                st.caption("Auditoría descargada. Botón bloqueado para evitar descargas duplicadas.")


def render_audit_tab():
    st.subheader("Auditoría de certificados")

    if not st.session_state.processed:
        st.markdown(
            """
            <div class="status-card">
                Aún no hay auditoría. Primero procesa uno o varios certificados PDF.
            </div>
            """,
            unsafe_allow_html=True,
        )
        return

    render_certificates_result_summary(show_downloads=False)


def render_downloads_tab():
    st.subheader("Descargas de certificados")

    if not st.session_state.processed:
        st.markdown(
            """
            <div class="status-card">
                Las descargas estarán disponibles después de procesar certificados PDF.
            </div>
            """,
            unsafe_allow_html=True,
        )
        return

    col_excel, col_audit = st.columns(2)

    with col_excel:
        st.download_button(
            label="Descargar Excel B2B completado",
            data=st.session_state.excel_bytes,
            file_name="B2B_TEMPLATE_GAMES_INTEGRATIONS_COMPLETADO.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
            disabled=st.session_state.excel_downloaded,
            on_click=mark_excel_downloaded,
        )
        if st.session_state.excel_downloaded:
            st.caption("Excel descargado. Botón bloqueado para evitar descargas duplicadas.")

    with col_audit:
        st.download_button(
            label="Descargar CSV de auditoría",
            data=st.session_state.audit_bytes,
            file_name="auditoria_certificados.csv",
            mime="text/csv",
            use_container_width=True,
            disabled=st.session_state.audit_downloaded,
            on_click=mark_audit_downloaded,
        )
        if st.session_state.audit_downloaded:
            st.caption("Auditoría descargada. Botón bloqueado para evitar descargas duplicadas.")


# =========================
# TAB MINCETUR
# =========================


def render_mincetur_tab():
    st.markdown(
        """
        <div class="step-row">
            <span class="mc-pill">1</span>Cargar resolución
            <span class="mc-pill">2</span>Analizar registros
            <span class="mc-pill">3</span>Validar tabla
            <span class="mc-pill">4</span>Descargar Excel
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.subheader("Resoluciones Directorales MINCETUR")
    st.write(
        "Carga una o varias Resoluciones Directorales de MINCETUR. La herramienta extraerá "
        "los registros homologados y generará un Excel con el formato regulatorio."
    )

    uploaded_files = st.file_uploader(
        "Arrastra o selecciona Resoluciones Directorales en PDF",
        type=["pdf"],
        accept_multiple_files=True,
        key=f"mincetur_uploader_{st.session_state.mincetur_uploader_key}",
    )

    if not uploaded_files and not st.session_state.mincetur_processed:
        st.markdown(
            """
            <div class="status-card">
                Carga una Resolución Directoral para extraer: N° correlativo, tipo de componente,
                nombre comercial, código de identificación del fabricante y número de registro MINCETUR.
            </div>
            """,
            unsafe_allow_html=True,
        )

    if uploaded_files:
        st.markdown(f"**Resoluciones cargadas:** {len(uploaded_files)}")
        with st.expander("Ver archivos cargados", expanded=True):
            for uploaded_file in uploaded_files:
                st.write(f"📄 {uploaded_file.name}")

        if st.button(
            "Analizar resoluciones MINCETUR",
            disabled=not uploaded_files,
            type="primary",
            use_container_width=True,
        ):
            with st.spinner("Analizando resoluciones..."):
                try:
                    with tempfile.TemporaryDirectory() as tmpdir:
                        tmpdir = Path(tmpdir)
                        pdf_dir = tmpdir / "resoluciones"
                        pdf_dir.mkdir(parents=True, exist_ok=True)

                        pdf_paths = []
                        for uploaded_file in uploaded_files:
                            pdf_path = pdf_dir / uploaded_file.name
                            with open(pdf_path, "wb") as file:
                                file.write(uploaded_file.getbuffer())
                            pdf_paths.append(pdf_path)

                        output_excel = tmpdir / "resoluciones_mincetur.xlsx"

                        rows, audit_rows = process_mincetur_resolutions(
                            pdfs=pdf_paths,
                            output_path=output_excel,
                        )

                        st.session_state.mincetur_rows = rows
                        st.session_state.mincetur_audit_rows = audit_rows
                        st.session_state.mincetur_excel_bytes = load_file_bytes(output_excel)
                        st.session_state.mincetur_processed = True
                        st.session_state.mincetur_downloaded = False
                        st.session_state.mincetur_uploader_key += 1

                    st.success("Resoluciones procesadas correctamente.")
                    st.rerun()

                except Exception as exc:
                    st.error(f"Ocurrió un error analizando las resoluciones: {exc}")

    if st.session_state.mincetur_processed:
        st.divider()
        st.markdown("### Resultado de Resoluciones MINCETUR")

        rows = st.session_state.mincetur_rows
        audit_rows = st.session_state.mincetur_audit_rows

        col1, col2 = st.columns(2)
        col1.metric("Registros extraídos", len(rows))
        col2.metric("PDFs analizados", len(audit_rows))

        if rows:
            df_rows = pd.DataFrame(rows)
            df_preview = df_rows[
                [
                    "tipo_componente",
                    "nombre_comercial",
                    "codigo_identificacion_fabricante",
                    "registro_mincetur",
                ]
            ].copy()

            df_preview.insert(0, "N° correlativo", range(1, len(df_preview) + 1))
            df_preview.columns = [
                "N° correlativo",
                "Tipo de componente",
                "Nombre comercial",
                "Código de identificación (fabricante)",
                "Número de registro del modelo otorgado por MINCETUR",
            ]

            st.dataframe(df_preview, use_container_width=True, hide_index=True)

            st.download_button(
                label="Descargar Excel de Resoluciones MINCETUR",
                data=st.session_state.mincetur_excel_bytes,
                file_name="resoluciones_mincetur.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
                disabled=st.session_state.mincetur_downloaded,
                on_click=mark_mincetur_downloaded,
            )

            if st.session_state.mincetur_downloaded:
                st.caption("Excel descargado. Botón bloqueado para evitar descargas duplicadas.")
        else:
            st.markdown(
                """
                <div class="warning-card">
                    No se encontraron registros homologados en los PDFs cargados.
                    Revisa que el documento corresponda a una Resolución Directoral MINCETUR con tabla del Artículo 1.
                </div>
                """,
                unsafe_allow_html=True,
            )

        with st.expander("Ver auditoría de resoluciones"):
            audit_df = pd.DataFrame(audit_rows)
            st.dataframe(style_audit_table(audit_df), use_container_width=True, hide_index=True)


# =========================
# APP
# =========================


def main():
    st.set_page_config(
        page_title=APP_TITLE,
        page_icon="🎰",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    init_session_state()
    render_css()
    render_sidebar()
    render_header()

    tab_certificates, tab_mincetur = st.tabs(
        [
            "📥 Certificados",
            "🏛️ Resoluciones MINCETUR",
        ]
    )

    with tab_certificates:
        render_certificates_tab()

    with tab_mincetur:
        render_mincetur_tab()


if __name__ == "__main__":
    main()
